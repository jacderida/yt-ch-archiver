import json
import os
import requests
import unicodedata

from datetime import datetime
from enum import Enum, auto
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from rich.console import Console
from rich.highlighter import RegexHighlighter
from rich.text import Text
from rich.theme import Theme


class WordHighlighter(RegexHighlighter):
    base_style = "hl."
    highlights = [r"(?P<word_unlisted>UNLISTED)",
                 "(?P<word_external>EXTERNAL)"]


def print_video(video_id, title, is_unlisted, is_private, is_external, is_downloaded):
    theme = Theme({"hl.word_unlisted": "blue", "hl.word_external": "yellow"})
    console = Console(highlighter=WordHighlighter(), theme=theme)
    msg = f"{video_id}: {title}"
    if is_unlisted or is_private or is_external:
        msg += " ["
        if is_unlisted:
            msg += "UNLISTED, "
        if is_private:
            msg += "PRIVATE, "
        if is_external:
            msg += "EXTERNAL"
        msg = msg.removesuffix(", ")
        msg += "]"
    if is_downloaded:
        console.print(msg, style="green")
    else:
        console.print(msg, style="red")


class Channel:
    def __init__(self, id, username, published_at, title, description):
        self.id = id
        self.username = username
        self.published_at = published_at
        self.title = title
        self.description = description
        self.video_count = 0
        self.video_size = ""
        self.video_duration = 0

    @staticmethod
    def from_channel_response_item(item):
        channel_id = item["id"]
        snippet = item["snippet"]
        published_at = snippet["publishedAt"]
        title = snippet["title"]
        if "customUrl" in snippet:
            username = snippet["customUrl"]
        else:
            username = title
        description = snippet["description"]
        return Channel(channel_id, username, published_at, title, description)

    @staticmethod
    def from_row(row):
        return Channel(row[0], row[1], row[2], row[3], row[4])


class Video:
    def __init__(self, id, title, channel_id, saved_path, is_unlisted, is_private):
        self.id = id
        self.title = title
        self.channel_id = channel_id
        self.saved_path = saved_path
        self.is_unlisted = is_unlisted
        self.is_private = is_private
        self.download_error = ""
        self.duration = ""
        self.resolution = ""

    def get_url(self):
        return f"http://www.youtube.com/watch?v={self.id}"

    @staticmethod
    def from_search_response_item(youtube, item, channel_id):
        id = item["id"]["videoId"]
        response = youtube.videos().list(part="snippet", id=id).execute()
        title = response["items"][0]["snippet"]["title"]
        print(f"Retrieved {id}: {title}")
        return Video(id, title, channel_id, "", False, False)

    @staticmethod
    def from_playlist_item(playlist_item):
        return Video(
            playlist_item.video_id,
            playlist_item.title,
            playlist_item.channel_id,
            "",
            playlist_item.is_unlisted,
            playlist_item.is_private,
        )

    @staticmethod
    def from_row(row):
        id = row[0]
        channel_id = row[1]
        title = row[2]
        saved_path = row[3]
        is_unlisted = row[4]
        is_private = row[5]
        video = Video(id, title, channel_id, saved_path, is_unlisted, is_private)
        if row[6]:
            video.download_error = row[6]
        if row[7]:
            video.duration = row[7]
        if row[8]:
            video.resolution = row[8]
        return video

    def print(self):
        is_downloaded = True if self.saved_path else False
        print_video(self.id, self.title, self.is_unlisted, self.is_private, False, is_downloaded)


class Playlist:
    def __init__(self, id, title, channel_id):
        self.id = id
        self.title = title
        self.channel_id = channel_id
        self.items = []

    class PlaylistItem:
        def __init__(
            self,
            id,
            video_id,
            channel_id,
            title,
            position,
            is_unlisted,
            is_private,
            is_external,
            is_deleted,
        ):
            self.id = id
            self.video_id = video_id
            self.channel_id = channel_id
            self.title = title
            self.position = position
            self.is_unlisted = True if is_unlisted == 1 else False
            self.is_private = True if is_private == 1 else False
            self.is_external = True if is_external == 1 else False
            self.is_deleted = True if is_deleted == 1 else False

        def print(self, video):
            if video:
                is_downloaded = True if video.saved_path else False
            else:
                is_downloaded = False
            print_video(
                self.video_id,
                self.title,
                self.is_unlisted,
                self.is_private,
                self.is_external,
                is_downloaded)

    def print_title(self):
        header = f"{self.title} ({len(self.items)} items)"
        console = Console()
        console.print(Text("=" * len(header), style="cyan"))
        console.print(f"{header}", style="cyan")
        console.print(Text("=" * len(header), style="cyan"))

    def add_item(
        self,
        id,
        video_id,
        channel_id,
        title,
        position,
        is_unlisted,
        is_private,
        is_external,
        is_deleted,
    ):
        item = self.PlaylistItem(
            id,
            video_id,
            channel_id,
            title,
            position,
            is_unlisted,
            is_private,
            is_external,
            is_deleted,
        )
        self.items.append(item)
        return item


class SyncReport:
    def __init__(self):
        self.start_time = datetime.now()
        self.videos_downloaded = {}
        self.failed_downloads = {}
        self.finish_time = None

    def _generate_report(self):
        report = []

        report.append("###############################################")
        report.append("#####                                     #####")
        report.append("#####             SYNC REPORT             #####")
        report.append("#####                                     #####")
        report.append("###############################################")
        report.append("Started: " + str(self.start_time))
        report.append("Finished: " + str(self.finish_time))
        
        report.append("\n################################")
        report.append("###### DOWNLOADED VIDEOS #######")
        report.append("################################")
        for channel, videos in self.videos_downloaded.items():
            small_banner = ""
            for _ in channel:
                small_banner += "="
            report.append(f"{small_banner}")
            report.append(f"{channel}")
            report.append(f"{small_banner}")
            if len(videos) > 0:
                for video in videos:
                    report.append(f"{video.id}: {video.title}")
            else:
                report.append("No videos downloaded")

        report.append("\n############################")
        report.append("###### FAILED VIDEOS #######")
        report.append("############################")
        for channel, videos in self.failed_downloads.items():
            small_banner = ""
            for _ in channel:
                small_banner += "="
            report.append(f"{small_banner}")
            report.append(f"{channel}")
            report.append(f"{small_banner}")
            if len(videos) > 0:
                for video in videos:
                    report.append(f"{video.id}: {video.title} -- {video.download_error}")
            else:
                report.append("No failed downloads")
        return "\n".join(report)

    def print(self):
        print(self._generate_report())

    def save(self, file_path):
        with open(file_path, 'w') as file:
            file.write(self._generate_report())


class VideoListSpreadsheet():
    def sanitize_string(self, s):
        # Remove control characters
        return ''.join(ch for ch in s if unicodedata.category(ch)[0]!="C")

    def generate_report(self, report_data, file_path):
        download_root_path = os.getenv("YT_CH_ARCHIVER_ROOT_PATH")
        if not download_root_path:
            raise Exception(
                "The YT_CH_ARCHIVER_ROOT_PATH environment variable must be set"
            )
        workbook = Workbook()
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        thumb_width = 150
        thumb_height = 150
        thumb_row_height = thumb_height * 0.76
        thumb_row_width = thumb_width / 8

        for channel_name, videos in report_data.items():
            channel_path = Path(download_root_path).joinpath(channel_name)
            print(f"Creating new sheet for {channel_name}")
            sheet = workbook.create_sheet(channel_name, 0)

            headers = ["", "ID", "Date", "Title", "Status", "Duration", "Resolution", "Description"]
            for col_num, header in enumerate(headers, 1):
                col_letter = chr(64 + col_num)
                sheet[f"{col_letter}1"] = header

            sheet_number = 2
            videos_len = len(videos)
            row_num = 2
            video_count = 1
            for video in videos:
                if row_num > 1000:
                    row_num = 2
                    print(f"Creating new sheet for {channel_name}")
                    sheet = workbook.create_sheet(f"{channel_name}{sheet_number}", 0)
                    sheet_number += 1
                    headers = ["", "ID", "Date", "Title", "Status", "Duration", "Resolution", "Description"]
                    for col_num, header in enumerate(headers, 1):
                        col_letter = chr(64 + col_num)
                        sheet[f"{col_letter}1"] = header

                print(f"Processing video {video_count} of {videos_len}: {video.id}")
                if video.saved_path:
                    title = video.title
                    duration = video.duration
                    resolution = video.resolution

                    info_path = channel_path.joinpath("info").joinpath(f"{video.id}.info.json")
                    if info_path.exists():
                        with open(info_path, "r") as info_file:
                            info = json.load(info_file)
                        upload_date = info["upload_date"]
                        upload_date = datetime.strptime(upload_date, "%Y%m%d").strftime("%Y-%m-%d")
                    else:
                        upload_date = "unknown"

                    description_path = channel_path.joinpath(
                        "description").joinpath(f"{video.id}.description")
                    with open(description_path, "r") as description_file:
                        description = description_file.read()

                    thumbnail_path = Path(video.saved_path).parent.parent.joinpath(
                        "thumbnail").joinpath(f"{video.id}.jpg")
                    if not thumbnail_path.exists():
                        raise Exception(f"Thumbnail not found at {thumbnail_path}")
                    thumbnail_image = Image(thumbnail_path)
                    thumbnail_image.width = thumb_width
                    thumbnail_image.height = thumb_height

                    sheet.row_dimensions[row_num].height = thumb_row_height
                    sheet.column_dimensions['A'].width = thumb_row_width
                
                    sheet.add_image(thumbnail_image, f"A{row_num}")
                elif video.is_private:
                    title = video.title
                    duration = "N/A"
                    resolution = "N/A"
                    description = ""
                    upload_date = ""
                else:
                    cleaned_error = "Video unavailable." + video.download_error.split("Video unavailable.", 1)[-1]
                    title = f"{video.title} -- {cleaned_error}"
                    duration = "N/A"
                    resolution = "N/A"
                    description = ""
                    upload_date = ""

                status = "PUBLIC"
                if video.is_private:
                    status = "PRIVATE"
                elif video.is_unlisted:
                    status = "UNLISTED"

                cell = sheet[f"B{row_num}"]
                cell.value = video.id
                cell.hyperlink = video.get_url()
                cell.style = "Hyperlink"
                sheet[f"C{row_num}"] = upload_date
                sheet[f"D{row_num}"] = self.sanitize_string(title)
                sheet[f"E{row_num}"] = status
                sheet[f"F{row_num}"] = duration
                sheet[f"G{row_num}"] = resolution
                sheet[f"H{row_num}"] = description

                if video.saved_path:
                    if video.is_unlisted:
                        sheet.cell(row=row_num, column=5).fill = blue_fill
                    else:
                        sheet.cell(row=row_num, column=5).fill = green_fill
                else:
                    for col_num in range(1, len(headers) + 1):
                        sheet.cell(row=row_num, column=col_num).fill = red_fill
                row_num += 1
                video_count += 1
        workbook.save(file_path)


class VideoDownloadReport:
    def __init__(self):
        self.start_time = datetime.now()
        self.videos_downloaded = []
        self.failed_downloads = []
        self.finish_time = None

    def _generate_report(self):
        report = []

        report.append("\n")
        report.append("###############################################")
        report.append("#####                                     #####")
        report.append("#####        VIDEO DOWNLOAD REPORT        #####")
        report.append("#####                                     #####")
        report.append("###############################################")
        report.append("Started: " + str(self.start_time))
        report.append("Finished: " + str(self.finish_time))

        report.append("\n################################")
        report.append("###### DOWNLOADED VIDEOS #######")
        report.append("################################")
        if len(self.videos_downloaded) > 0:
            for video in self.videos_downloaded:
                report.append(f"{video.id}: {video.title}")
        else:
            report.append("No videos downloaded")

        report.append("\n############################")
        report.append("###### FAILED VIDEOS #######")
        report.append("############################")
        if len(self.failed_downloads) > 0:
            for video in self.failed_downloads:
                report.append(f"{video.id}: {video.title} -- {video.download_error}")
        else:
            report.append("No failed downloads")
        return "\n".join(report)

    def add_data_sources(self, videos_downloaded, failed_downloads):
        self.videos_downloaded.extend(videos_downloaded)
        self.failed_downloads.extend(failed_downloads)

    def mark_finished(self):
        self.finish_time = datetime.now()

    def print(self):
        print(self._generate_report())

    def save(self, file_path):
        with open(file_path, 'w') as file:
            file.write(self._generate_report())
